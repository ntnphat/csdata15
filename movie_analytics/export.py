from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from movie_analytics import charts, metrics
from movie_analytics.etl import data_quality_report

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN = Side(style="thin", color="D9DDE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_COLUMNS = {"median_budget", "median_gross", "total_gross", "budget_real", "gross_real",
                 "profit_real", "budget", "gross"}
RATIO_COLUMNS = {"median_multiple", "median_roi", "multiple", "roi", "spread",
                 "p25_multiple", "p75_multiple"}
PERCENT_COLUMNS = {"hit_rate", "loss_rate", "us_share"}

DETAIL_COLUMNS = ["name", "year", "genre", "rating_group", "country", "director", "star",
                  "company", "season", "runtime", "score", "votes", "budget_real",
                  "gross_real", "profit_real", "multiple"]


def _style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, df: pd.DataFrame, start_col: int = 1, max_width: int = 32) -> None:
    for i, column in enumerate(df.columns, start=start_col):
        longest = max([len(str(column))] + [len(str(v)) for v in df[column].head(200)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_width, max(11, longest + 2))


def _write_table(ws, df: pd.DataFrame, title: str, start_row: int = 1) -> int:
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    header_row = start_row + 1

    for j, column in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=j, value=str(column))
    _style_header(ws, header_row, len(df.columns))

    for i, (_, record) in enumerate(df.iterrows(), start=header_row + 1):
        for j, column in enumerate(df.columns, start=1):
            value = record[column]
            cell = ws.cell(row=i, column=j, value=None if pd.isna(value) else value)
            cell.border = BORDER
            if column in MONEY_COLUMNS:
                cell.number_format = '#,##0'
            elif column in RATIO_COLUMNS:
                cell.number_format = '0.00"x"'
            elif column in PERCENT_COLUMNS:
                cell.number_format = '0.0"%"'
            elif isinstance(value, float):
                cell.number_format = '#,##0.00'

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, df)
    return header_row + len(df) + 3


def _write_kpi_sheet(ws, df: pd.DataFrame, summary_text: str = "") -> None:
    kpi = metrics.kpi_summary(df)
    rows = [
        ("Số phim trong phạm vi báo cáo / Titles", f"{kpi['titles']:,}"),
        ("Số phim đủ dữ liệu tài chính / With financials", f"{kpi['titles_with_financials']:,}"),
        ("Độ phủ dữ liệu tài chính / Coverage", f"{kpi['coverage_pct']:.1f}%"),
        ("Giai đoạn / Period", f"{kpi['year_min']} - {kpi['year_max']}"),
        ("Tổng ngân sách / Total budget (real 2020 USD)", f"${kpi['total_budget']:,.0f}"),
        ("Tổng doanh thu / Total gross (real 2020 USD)", f"${kpi['total_gross']:,.0f}"),
        ("Ngân sách trung vị / Median budget", f"${kpi['median_budget']:,.0f}"),
        ("Doanh thu trung vị / Median gross", f"${kpi['median_gross']:,.0f}"),
        ("Bội số thu hồi vốn trung vị / Median multiple", f"{kpi['median_multiple']:.2f}x"),
        ("Tỷ lệ đạt hòa vốn 2.0x / Hit rate", f"{kpi['hit_rate']:.1f}%"),
        ("Tỷ lệ gross > budget (chưa tính P&A)", f"{kpi['naive_hit_rate']:.1f}%"),
        ("Điểm IMDb trung bình / Mean score", f"{kpi['mean_score']:.2f}"),
    ]
    table = pd.DataFrame(rows, columns=["Chỉ số / Metric", "Giá trị / Value"])
    next_row = _write_table(ws, table, "BÁO CÁO HIỆU QUẢ TÀI CHÍNH NGÀNH ĐIỆN ẢNH 1980-2020")

    if summary_text:
        ws.cell(row=next_row, column=1, value="Tóm tắt điều hành / Executive summary").font = TITLE_FONT
        cell = ws.cell(row=next_row + 1, column=1, value=summary_text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 18, end_column=6)


def _insert_chart(ws, fig, anchor: str) -> None:
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=130, bbox_inches="tight")
    buffer.seek(0)
    ws.add_image(XLImage(buffer), anchor)


def build_excel_report(df: pd.DataFrame, raw: pd.DataFrame = None,
                       summary_text: str = "", include_charts: bool = True,
                       detail_rows: int = 2000) -> BytesIO:
    wb = Workbook()

    _write_kpi_sheet(wb.active, df, summary_text)
    wb.active.title = "1. Tóm tắt"

    if raw is not None:
        _write_table(wb.create_sheet("2. Chất lượng DL"), data_quality_report(raw),
                     "Chất lượng dữ liệu nguồn / Source data quality")

    sheets = {
        "3. Thể loại": (metrics.risk_return_table(df, "genre", 20), "Rủi ro - lợi nhuận theo thể loại"),
        "4. Phân loại tuổi": (metrics.group_performance(df, "rating_group", 10), "Hiệu quả theo nhãn phân loại"),
        "5. Tầng ngân sách": (metrics.group_performance(df, "budget_tier", 10), "Hiệu quả theo tầng ngân sách"),
        "6. Mùa phát hành": (metrics.group_performance(df, "season", 10), "Hiệu quả theo mùa phát hành"),
        "7. Xu hướng năm": (metrics.yearly_trend(df), "Diễn biến theo năm"),
        "8. Pareto": (metrics.pareto_concentration(df), "Mức độ tập trung lợi nhuận"),
        "9. Điểm IMDb": (metrics.score_vs_money(df), "Điểm IMDb và hiệu quả tài chính"),
    }
    for sheet_name, (table, title) in sheets.items():
        if not table.empty:
            _write_table(wb.create_sheet(sheet_name), table, title)

    ws_rank = wb.create_sheet("10. Xếp hạng")
    cursor = 1
    for column, label in [("director", "Đạo diễn"), ("star", "Diễn viên chính"), ("company", "Hãng phim")]:
        table = metrics.top_entities(df, column, top_n=15)
        if not table.empty:
            cursor = _write_table(ws_rank, table, f"Top 15 {label} theo bội số thu hồi vốn", cursor)

    detail = df[[c for c in DETAIL_COLUMNS if c in df.columns]].copy()
    detail = detail.sort_values("gross_real", ascending=False).head(detail_rows)
    _write_table(wb.create_sheet("11. Dữ liệu chi tiết"), detail,
                 f"Chi tiết {len(detail):,} phim doanh thu cao nhất trong phạm vi lọc")

    if include_charts:
        ws_chart = wb.create_sheet("12. Biểu đồ")
        ws_chart.cell(row=1, column=1, value="Biểu đồ tổng hợp / Key charts").font = TITLE_FONT
        figures = [
            (charts.risk_return_scatter(metrics.risk_return_table(df, "genre", 20)), "A3"),
            (charts.pareto_chart(metrics.pareto_concentration(df)), "A30"),
            (charts.trend_lines(metrics.yearly_trend(df)), "A57"),
        ]
        for fig, anchor in figures:
            _insert_chart(ws_chart, fig, anchor)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
